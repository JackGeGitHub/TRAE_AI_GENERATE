
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# 创建工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "工程源码清单"

# 定义表头样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# 定义数据
data = [
    ["序号", "文件名称", "相对路径", "文件类型", "代码行数", "功能描述"],
    ["1", "MainActivity.java", "app/src/main/java/com/capturescreen/app/MainActivity.java", "Java源文件", 150, "应用主界面，处理权限申请和服务启停"],
    ["2", "CaptureScreenService.java", "app/src/main/java/com/capturescreen/app/CaptureScreenService.java", "Java源文件", 515, "核心服务，管理截屏流程、悬浮窗和MediaProjection"],
    ["3", "ScreenCapturer.java", "app/src/main/java/com/capturescreen/app/ScreenCapturer.java", "Java源文件", 121, "屏幕捕获工具类，从ImageReader获取图像并裁剪"],
    ["4", "CaptureRegionView.java", "app/src/main/java/com/capturescreen/app/CaptureRegionView.java", "Java源文件", 266, "截屏区域视图，支持双指缩放和长按拖拽"],
    ["5", "ChatBottomSheetActivity.java", "app/src/main/java/com/capturescreen/app/ChatBottomSheetActivity.java", "Java源文件", 156, "底部弹窗，显示截图并保存到相册"],
    ["6", "HintView.java", "app/src/main/java/com/capturescreen/app/HintView.java", "Java源文件", 91, "提示视图，显示拖拽分享提示"],
    ["7", "AndroidManifest.xml", "app/src/main/AndroidManifest.xml", "配置文件", 44, "应用清单配置，权限声明和组件注册"],
    ["8", "activity_main.xml", "app/src/main/res/layout/activity_main.xml", "布局文件", 34, "主界面布局"],
    ["9", "activity_chat_bottom_sheet.xml", "app/src/main/res/layout/activity_chat_bottom_sheet.xml", "布局文件", 31, "底部弹窗布局"],
    ["10", "overlay_view.xml", "app/src/main/res/layout/overlay_view.xml", "布局文件", 5, "悬浮窗布局"],
    ["11", "strings.xml", "app/src/main/res/values/strings.xml", "资源文件", 16, "字符串资源"],
    ["12", "colors.xml", "app/src/main/res/values/colors.xml", "资源文件", 12, "颜色资源"],
    ["13", "themes.xml", "app/src/main/res/values/themes.xml", "资源文件", 14, "主题样式"],
    ["14", "build.gradle", "app/build.gradle", "Gradle配置", 38, "应用模块Gradle配置"],
    ["15", "build.gradle", "build.gradle", "Gradle配置", 3, "项目根Gradle配置"],
    ["16", "settings.gradle", "settings.gradle", "Gradle配置", 23, "项目设置配置"],
    ["17", "gradle.properties", "gradle.properties", "配置文件", 3, "Gradle属性配置"],
    ["18", "proguard-rules.pro", "app/proguard-rules.pro", "配置文件", 1, "ProGuard混淆规则"],
    ["19", "README.md", "README.md", "文档", 0, "项目说明文档"]
]

# 写入数据
for row_idx, row_data in enumerate(data, 1):
    for col_idx, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
        # 表头样式
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

# 调整列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 40

# 保存文件
output_path = r"d:\Workspace\Android\Codes\TRAE_AI_GENERATE\CaptureScreen\doc\CaptureScreenCodes.xlsx"
wb.save(output_path)
print(f"Excel文件已成功生成: {output_path}")
