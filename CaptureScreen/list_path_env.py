
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 获取PATH环境变量
path_env = os.environ.get('PATH', '')
path_dirs = path_env.split(os.pathsep)

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PATH环境变量列表"

# 定义表头样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# 写入表头
headers = ["序号", "路径", "是否存在"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 写入数据
for row_idx, directory in enumerate(path_dirs, 2):
    exists = "是" if (directory and os.path.exists(directory)) else "否"
    ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
    ws.cell(row=row_idx, column=2, value=directory)  # 路径
    ws.cell(row=row_idx, column=3, value=exists)  # 是否存在

# 调整列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 100
ws.column_dimensions['C'].width = 10

# 保存文件
output_path = r"d:\Workspace\Android\Codes\TRAE_AI_GENERATE\CaptureScreen\doc\ListPowerShellPath.xlsx"
wb.save(output_path)
wb.close()

print(f"PATH环境变量包含 {len(path_dirs)} 个路径")
print(f"结果已保存到: {output_path}")
