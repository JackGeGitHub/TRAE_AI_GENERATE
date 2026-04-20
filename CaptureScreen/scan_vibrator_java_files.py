
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 目标目录
target_dir = r"D:\Workspace\AndroidSdk\sources\android-35\android\os\vibrator"

# 收集所有Java文件的绝对路径
java_files = []

# 遍历目录及其子目录
for root, dirs, files in os.walk(target_dir):
    for file in files:
        if file.endswith('.java'):
            file_path = os.path.join(root, file)
            java_files.append(file_path)

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Java文件清单"

# 定义表头样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# 写入表头
headers = ["序号", "Java文件绝对路径"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 写入数据
for row_idx, file_path in enumerate(java_files, 2):
    ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 序号
    ws.cell(row=row_idx, column=2, value=file_path)  # 绝对路径

# 调整列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 100

# 保存文件
output_path = r"d:\Workspace\Android\Codes\TRAE_AI_GENERATE\CaptureScreen\doc\SummaryJavaFile.xlsx"
wb.save(output_path)
wb.close()

print(f"已找到 {len(java_files)} 个Java文件")
print(f"结果已保存到: {output_path}")
